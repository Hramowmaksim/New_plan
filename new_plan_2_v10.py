import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl import Workbook
import re
INVALID_CHARS = re.compile(r'[,;*?<>|":\\/]')   # запрещённые символы
NON_ASCII = re.compile(r'[^\x00-\x7F]')          # всё, что вне ASCII (иероглифы, эмодзи)

# ------------------------------------------------------------------
# Габариты 40-футового контейнера (мм)
L40, W40, H40 = 12_032, 2_352, 2_395
SCALE = 0.07                       # px / mm
# ------------------------------------------------------------------

def validate_name(name: str) -> str:
    """
    Удаляет пробелы по краям, проверяет запрещённые символы.
    Возвращает имя или выбрасывает ValueError.
    """
    name = name.strip()
    if not name:
        raise ValueError("Название не может быть пустым")
    if INVALID_CHARS.search(name):
        raise ValueError("Название содержит запрещённые символы: , ; * ? < > | \" : \\ /")
    if NON_ASCII.search(name):
        raise ValueError("Название содержит иероглифы или спецсимволы")
    return name


MAX_WEIGHT_KG = 28_000
MAX_VOLUME_M3 = 33.2   # 20-футовый контейнер


# ------------------------------------------------------------------
class ContainerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Размещение груза в контейнере")
        self.drag_instance = None   # текущий перетаскиваемый экземпляр

        # ---- добавьте эту строку ----
        self.entries = {}
        # ----------------------------

        # габариты контейнера (мм)
        self.L, self.W, self.H = 12032, 2350, 2381
        self.tw, self.th = int(self.L * SCALE), int(self.W * SCALE)
        self.sw, self.sh = int(self.L * SCALE), int(self.H * SCALE)
        self.fw, self.fh = int(self.W * SCALE), int(self.H * SCALE)

        self.cargos = []          # все грузы
        self.counter = 1          # для генерации id
        self.colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
                       "#9467bd", "#8c564b", "#e377c2", "#7f7f7f",
                       "#bcbd22", "#17becf"]

        self.build_gui()
        self.draw_containers()
        self.root.bind_all("<KeyPress-r>", self.on_rotate_key)  # <-- сюда

    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    def build_gui(self):
        # левая панель
        left = tk.Frame(self.root, width=300)
        left.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)
        left.pack_propagate(False)

        # ---------------- форма ввода ----------------
        frm = tk.LabelFrame(left, text="Параметры груза")
        frm.pack(fill=tk.X, pady=(0, 10))

        # контейнер для двух колонок
        cols = tk.Frame(frm)
        cols.pack(pady=5)

        # колонка 1
        col1 = tk.Frame(cols)
        col1.pack(side=tk.LEFT, padx=5)
        for txt in ("Название:", "Количество:", "Вес (кг):"):
            tk.Label(col1, text=txt, width=12, anchor='w').pack()
            e = tk.Entry(col1, width=15)
            e.pack()

        # колонка 2
        col2 = tk.Frame(cols)
        col2.pack(side=tk.LEFT, padx=5)
        for txt in ("Длина (мм):", "Ширина (мм):", "Высота (мм):"):
            tk.Label(col2, text=txt, width=12, anchor='w').pack()
            e = tk.Entry(col2, width=15)
            e.pack()

        # ---------------- кнопки под формой ----------------
        # ---------------- кнопки 2×3 ----------------
        btns = tk.Frame(left)
        btns.pack(fill=tk.X, pady=5)

        # строка 1
        row1 = tk.Frame(btns)
        row1.pack()
        tk.Button(row1, text="Добавить",  command=self.add_to_table,    width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row1, text="Экспорт",   command=self.export_xlsx,     width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row1, text="Импорт",    command=self.import_xlsx,     width=12).pack(side=tk.LEFT, padx=3)

        # строка 2
        row2 = tk.Frame(btns)
        row2.pack()
        tk.Button(row2, text="Разместить", command=self.place_selected, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row2, text="Удалить",    command=self.delete_selected, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row2, text="Очистить",   command=self.clear_canvas,    width=12).pack(side=tk.LEFT, padx=3)

        # ---------------- дальше всё как было ----------------
        self.status_lbl = tk.Label(left, text="Вес: 0 / 28 000 кг   Объём: 0 / 33 м³")
        self.status_lbl.pack(pady=5)

        self.tree = ttk.Treeview(left,
                                 columns=("name", "qty", "l", "w", "h", "wt", "placed", "left"),
                                 show="headings", height=10)
        self.tree.heading("name", text="Название")
        self.tree.heading("qty", text="Кол-во")
        self.tree.heading("l", text="Длина")
        self.tree.heading("w", text="Ширина")
        self.tree.heading("h", text="Высота")
        self.tree.heading("wt", text="Вес")
        self.tree.heading("placed", text="Размещено")
        self.tree.heading("left", text="Осталось")
        for col in ("name", "qty", "l", "w", "h", "wt", "placed", "left"):
            self.tree.column(col, width=80, anchor="center")
        self.tree.pack(fill=tk.BOTH, expand=True)


      # --------------------------------------------------------------
        # ПРАВАЯ ПАНЕЛЬ  (Canvas)
        right = tk.Frame(self.root)
        right.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # ---------- 1. Вид сбоку (самый верх) ----------
        side_box = tk.Frame(right)
        side_box.pack(anchor='w', pady=(0, 5))
        self.side_canvas = tk.Canvas(side_box,
                                     width=self.sw + 20,
                                     height=self.sh + 20,
                                     bg="gray95")
        self.side_canvas.pack()
        tk.Label(side_box, text="Вид сбоку").pack()

        # ---------- 2. Вид сверху (середина) ----------
        top_box = tk.Frame(right)
        top_box.pack(anchor='w', pady=5)
        self.top_canvas = tk.Canvas(top_box,
                                    width=self.tw + 20,
                                    height=self.th + 20,
                                    bg="gray95")
        self.top_canvas.pack()
        tk.Label(top_box, text="Вид сверху").pack()

        # ---------- 3. Вид спереди (самый низ, влево) ----------
        front_box = tk.Frame(right)
        front_box.pack(anchor='w', pady=(5, 0))
        self.front_canvas = tk.Canvas(front_box,
                                      width=self.fw + 20,
                                      height=self.fh + 20,
                                      bg="gray95")
        self.front_canvas.pack()
        tk.Label(front_box, text="Вид спереди").pack()
        
        # ⬇⬇⬇  ВСТАВЛЯЕМ СЮДА ⬇⬇⬇
        self.top_tooltip  = Tooltip(self.top_canvas, self)
        self.side_tooltip = Tooltip(self.side_canvas, self)
        self.front_tooltip = Tooltip(self.front_canvas, self)
        # ⬆⬆⬆  КОНЕЦ ВСТАВКИ  ⬆⬆⬆

        # bindings
        self.top_canvas.bind("<Button-1>", self.start_drag)
        self.top_canvas.bind("<B1-Motion>", self.do_drag)
        self.top_canvas.bind("<ButtonRelease-1>", self.end_drag)

        # двойной клик по прямоугольнику убирает груз
        for c in (self.top_canvas, self.side_canvas, self.front_canvas):
            c.bind("<Double-1>", self.remove_from_canvas)

    # ------------------------------------------------------------------
    def draw_containers(self):
        # Вычисляем пиксельные размеры
        tw = int(L40 * SCALE)
        th = int(W40 * SCALE)
        sw = int(L40 * SCALE)
        sh = int(H40 * SCALE)
        fw = int(W40 * SCALE)
        fh = int(H40 * SCALE)

        # Сохраним для удобства
        self.tw, self.th = tw, th
        self.sw, self.sh = sw, sh
        self.fw, self.fh = fw, fh

        # Сетка шаг 500 мм
        grid_step = 500
        grid_color = "#dddddd"

        for canv, w_px, h_px in ((self.top_canvas, tw, th),
                                 (self.side_canvas, sw, sh),
                                 (self.front_canvas, fw, fh)):
            # контур контейнера
            canv.create_rectangle(10, 10, 10 + w_px, 10 + h_px,
                                  width=2, outline="black", fill="white")

            # вертикальные линии
            x = 10 + grid_step * SCALE
            while x < 10 + w_px:
                canv.create_line(x, 10, x, 10 + h_px, fill=grid_color)
                x += grid_step * SCALE

            # горизонтальные линии
            y = 10 + grid_step * SCALE
            while y < 10 + h_px:
                canv.create_line(10, y, 10 + w_px, y, fill=grid_color)
                y += grid_step * SCALE
                
        # в конце build_gui
     
    # ------------------------------------------------------------------
    # CRUD-методы
    # ------------------------------------------------------------------
    def add_to_table(self):
        try:
            name = validate_name(self.entries["Название:"].get())
            qty  = int(self.entries["Количество:"].get())
            l = int(self.entries["Длина (мм):"].get())
            w = int(self.entries["Ширина (мм):"].get())
            h = int(self.entries["Высота (мм):"].get())
            wt = int(self.entries["Вес (кг):"].get())

            if qty <= 0 or min(l, w, h, wt) <= 0:
                raise ValueError
            if l > self.L or w > self.W or h > self.H:
                messagebox.showerror("Ошибка", "Габариты превышают контейнер")
                return

            cargo = {
                'id': self.counter,
                'name': name,
                'qty': qty,
                'length': l, 'width': w, 'height': h, 'weight': wt,
                'placed': 0          # счётчик размещённых
            }
            self.cargos.append(cargo)
            self.tree.insert("", "end", iid=str(self.counter),
                             values=(name, qty, l, w, h, wt, 0, qty))
            self.counter += 1
            self.clear_entries()
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте ввод")

    def refresh_tree_row(self, cargo):
        left = cargo['qty'] - cargo['placed']
        self.tree.item(str(cargo['id']),
                       values=(cargo['name'], cargo['qty'],
                               cargo['length'], cargo['width'], cargo['height'],
                               cargo['weight'], cargo['placed'], left))

    # ------------------------------------------------------------------
    # метод удаления груза из таблицы для кнопки удалить
    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз для удаления")
            return

        cid = int(sel[0])
        cargo = next((c for c in self.cargos if c['id'] == cid), None)
        if not cargo:
            return

        # 1. удаляем все экземпляры с холстов
        for inst in cargo.get('instances', []):
            for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
                canv.delete(inst['top_id'])
                canv.delete(inst['side_id'])
                canv.delete(inst['front_id'])

        # 2. удаляем строку из Treeview
        self.tree.delete(sel[0])

        # 3. удаляем из списка cargos
        self.cargos = [c for c in self.cargos if c['id'] != cid]

        # 4. обновляем статус
        self.update_status()

    def place_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз в таблице")
            return
        cid = int(sel[0])
        cargo = next(c for c in self.cargos if c['id'] == cid)

        # стартовые координаты — по центру контейнера
        x = (self.L - cargo['length']) // 2
        y = (self.W - cargo['width'])  // 2

        # максимальная высота в этой (x,y) области
        max_z = 0
        for c in self.cargos:
            for inst in c.get('instances', []):
                # пересекаются ли проекции на плоскость XY?
                if (x < inst['x'] + inst['length'] and x + cargo['length'] > inst['x'] and
                    y < inst['y'] + inst['width']  and y + cargo['width']  > inst['y']):
                    max_z = max(max_z, inst['z'] + inst['height'])

        z = max_z
        color = random.choice(self.colors)
        l, w, h = cargo['length'], cargo['width'], cargo['height']

        top_id = self.top_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE,
            fill=color, outline="black",
            tags=("cargo", f"{cargo['id']}_{cargo['placed']}"))

        side_id = self.side_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE,
            fill=color, outline="black",
            tags=("cargo", f"{cargo['id']}_{cargo['placed']}"))

        front_id = self.front_canvas.create_rectangle(
            10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE,
            fill=color, outline="black",
            tags=("cargo", f"{cargo['id']}_{cargo['placed']}"))

        cargo.setdefault('instances', []).append({
            'x': x, 'y': y, 'z': z,
            'top_id': top_id, 'side_id': side_id, 'front_id': front_id,
            'length': l, 'width': w, 'height': h,
            'name': cargo['name'], 'weight': cargo['weight']
        })
        cargo['placed'] += 1
        self.refresh_tree_row(cargo)
        self.update_status()


    # ------------------------------------------------------------------
    # удаление из контейнера
    # ------------------------------------------------------------------
    def remove_from_canvas(self, evt):
        canvas = evt.widget
        items  = canvas.find_withtag("current")
        if not items:
            return
        clicked_id = items[0]

        # ищем экземпляр по любому из его id
        for c in self.cargos:
            for idx, inst in enumerate(c.get('instances', [])):
                if clicked_id in (inst['top_id'], inst['side_id'], inst['front_id']):
                    # удаляем с холстов
                    for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
                        canv.delete(inst['top_id'])
                        canv.delete(inst['side_id'])
                        canv.delete(inst['front_id'])
                    # удаляем из списка экземпляров
                    c['instances'].pop(idx)
                    c['placed'] -= 1
                    self.refresh_tree_row(c)
                    self.update_status()
                    return
                
    # ------------------------------------------------------------------
    # столкновение
    # ------------------------------------------------------------------
    def collision_3d(self, x, y, z, l, w, h, exclude_instance=None):
        """True, если новый AABB пересекается с уже размещёнными.
        exclude_instance – экземпляр, который не учитываем
                            (например, двигаемый/вращаемый сам на себя).
        """
        for c in self.cargos:
            for inst in c.get('instances', []):
                if inst is exclude_instance:          # <-- ключевая строка
                    continue
                if (x < inst['x'] + inst['length'] and x + l > inst['x'] and
                    y < inst['y'] + inst['width']  and y + w > inst['y'] and
                    z < inst['z'] + inst['height'] and z + h > inst['z']):
                    return True
        return False

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def export_xlsx(self):
        if not self.cargos:
            messagebox.showwarning("Нет данных", "Нечего экспортировать")
            return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Грузы"
        ws.append(["№", "Название", "Количество", "Длина (мм)", "Ширина (мм)", "Высота (мм)", "Вес (кг)"])
        for c in self.cargos:
            ws.append([c['id'], c['name'], c['qty'],
                       c['length'], c['width'], c['height'], c['weight']])
        wb.save(f)
        messagebox.showinfo("Готово", f"Сохранено в {f}")

    def import_xlsx(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 7:
                    _, name, qty, l, w, h, wt = row[:7]
                    if None in (name, qty, l, w, h, wt):
                        continue
                    cargo = {
                        'id': self.counter,
                        'name': str(name),
                        'qty': int(qty),
                        'length': int(l), 'width': int(w), 'height': int(h),
                        'weight': int(wt),
                        'placed': 0
                    }
                    self.cargos.append(cargo)
                    self.tree.insert("", "end", iid=str(self.counter),
                                     values=(name, qty, l, w, h, wt, 0, qty))
                    self.counter += 1
        except Exception as e:
            messagebox.showerror("Ошибка импорта", str(e))

    # ------------------------------------------------------------------
    # drag-and-drop (только по уже размещённым)
    # ------------------------------------------------------------------
    def start_drag(self, evt):
        items = self.top_canvas.find_withtag("current")
        if not items:
            self.drag_instance = None
            return
        clicked_id = items[0]
        # ищем экземпляр, чей top_id == clicked_id
        for c in self.cargos:
            if not c.get('instances'):
                continue
            for inst in c['instances']:
                if inst['top_id'] == clicked_id:
                    self.drag_instance = inst
                    self.drag_start = (evt.x, evt.y)
                    return
        self.drag_instance = None

    def do_drag(self, evt):
        if not self.drag_instance:
            return
        dx = int((evt.x - self.drag_start[0]) / SCALE)
        dy = int((evt.y - self.drag_start[1]) / SCALE)

        l = self.drag_instance['length']
        w = self.drag_instance['width']
        h = self.drag_instance['height']

        new_x = max(0, min(self.L - l, self.drag_instance['x'] + dx))
        new_y = max(0, min(self.W - w, self.drag_instance['y'] - dy))
        new_z = self.drag_instance['z']

        if self.collision_3d(new_x, new_y, new_z, l, w, h,
                            exclude_instance=self.drag_instance):
            return

        self.move_instance_to(self.drag_instance, new_x, new_y, new_z)
        self.drag_start = (evt.x, evt.y)

        
    def end_drag(self, evt):
        if not self.drag_instance:
            return

        # новые координаты по X,Y
        dx = int((evt.x - self.drag_start[0]) / SCALE)
        dy = int((evt.y - self.drag_start[1]) / SCALE)

        l = self.drag_instance['length']
        w = self.drag_instance['width']

        new_x = max(0, min(self.L - l, self.drag_instance['x'] + dx))
        new_y = max(0, min(self.W - w, self.drag_instance['y'] - dy))

        # «падаем» на самый верх в этой точке
        new_z = self.drop_z(new_x, new_y, l, w, exclude_instance=self.drag_instance)

        # проверка коллизии (исключаем самого себя)
        if self.collision_3d(new_x, new_y, new_z, l, w,
                             self.drag_instance['height'],
                             exclude_instance=self.drag_instance):
            return   # на всякий случай, если вдруг пересечение

        self.move_instance_to(self.drag_instance, new_x, new_y, new_z)
        self.drag_instance = None



    def move_instance_to(self, inst, x, y, z):
        l, w, h = inst['length'], inst['width'], inst['height']
        inst['x'], inst['y'], inst['z'] = x, y, z

        self.top_canvas.coords(inst['top_id'],
            10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE)

        self.side_canvas.coords(inst['side_id'],
            10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE)

        self.front_canvas.coords(inst['front_id'],
            10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE)



    # ------------------------------------------------------------------
    # misc
    # ------------------------------------------------------------------
    def update_status(self):
        placed = [c for c in self.cargos if c.get('placed')]
        total_w = sum(c['weight'] for c in placed)
        total_v = sum(c['length'] * c['width'] * c['height'] for c in placed) / 1e9
        self.status_lbl.config(text=f"Вес: {total_w} / {MAX_WEIGHT_KG} кг   "
                                    f"Объём: {total_v:.2f} / {MAX_VOLUME_M3} м³")
        
    # ------------------------------------------------------------------
    # Поворот на 90°
    # ------------------------------------------------------------------
    def on_rotate_key(self, evt):
        items = self.top_canvas.find_withtag("current")
        if not items:
            return
        clicked_id = items[0]

        for c in self.cargos:
            for inst in c.get('instances', []):
                if inst['top_id'] == clicked_id:
                    new_l = inst['width']
                    new_w = inst['length']
                    h      = inst['height']

                    if (inst['x'] + new_l > self.L) or (inst['y'] + new_w > self.W):
                        messagebox.showerror("Ошибка", "Поворот выходит за границы")
                        return

                    # проверяем пересечение с остальными
                    if self.collision_3d(inst['x'], inst['y'], inst['z'],
                                        new_l, new_w, h,
                                        exclude_instance=inst):   # <-- вот здесь
                        messagebox.showerror("Ошибка", "Поворот приведёт к пересечению")
                        return

                    # всё ок – меняем размеры и перерисовываем
                    inst['length'], inst['width'] = new_l, new_w
                    self.move_instance_to(inst, inst['x'], inst['y'], inst['z'])
                    return

    def redraw_rotated(self, c):
        l, w, h = c['length'], c['width'], c['height']
        x, y, z = c['x'], c['y'], c['z']

        self.top_canvas.coords(c['top_id'],
                               10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
                               10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE)

        self.side_canvas.coords(c['side_id'],
                                10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
                                10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE)

        self.front_canvas.coords(c['front_id'],
                                 10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
                                 10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE)

    # --------------------------------------------------
    def drop_z(self, x, y, l, w, exclude_instance=None):
        """Возвращает высоту, на которой должен лечь груз (x,y,l,w)."""
        max_z = 0
        for c in self.cargos:
            for inst in c.get('instances', []):
                if inst is exclude_instance:      # не учитываем самого себя
                    continue
                if (x < inst['x'] + inst['length'] and x + l > inst['x'] and
                    y < inst['y'] + inst['width']  and y + w > inst['y']):
                    max_z = max(max_z, inst['z'] + inst['height'])
        return max_z

    def clear_canvas(self):
        """Удалить все прямоугольники с холстов, но оставить записи в таблице."""
        for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
            canv.delete("cargo")
        for c in self.cargos:
            c['placed'] = 0
            c['instances'].clear()
            self.refresh_tree_row(c)
        self.update_status()

    def clear_entries(self):
        for e in self.entries.values():
            e.delete(0, tk.END)


# ------------------------------------------------------------------

class Tooltip:
    """Простой tooltip для Canvas"""
    def __init__(self, canvas, app):
        self.canvas = canvas
        self.app = app
        self.tooltip = None
        canvas.bind("<Motion>", self.on_motion)
        canvas.bind("<Leave>", self.hide)

    # -------------------------------------------------
    def on_motion(self, event):
        items = event.widget.find_withtag("current")
        if not items:
            self.hide()
            return
        clicked_id = items[0]

        # ищем именно экземпляр, у которого один из canvas-ids совпал
        cargo = inst = None
        for c in self.app.cargos:
            for ins in c.get('instances', []):
                if clicked_id in (ins['top_id'], ins['side_id'], ins['front_id']):
                    cargo = c
                    inst  = ins
                    break
            if cargo:
                break

        if cargo:
            text = f"{cargo['name']}  {inst['weight']} кг"
            self.show(text, event.x_root, event.y_root)
        else:
            self.hide()

    # -------------------------------------------------
    def show(self, text, x, y):
        if self.tooltip is None:
            self.tooltip = tk.Toplevel(self.canvas)
            self.tooltip.wm_overrideredirect(True)
            self.label = tk.Label(self.tooltip, text=text,
                                  bg="yellow", relief="solid", bd=1)
            self.label.pack()
        else:
            self.label.config(text=text)
        self.tooltip.wm_geometry(f"+{x + 10}+{y + 10}")

    # -------------------------------------------------
    def hide(self, *args):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

if __name__ == "__main__":
    root = tk.Tk()
    ContainerApp(root)
    root.mainloop()

#крайняя рабочая версия приложения. Добавлены типсы. Добавлено тримирование названий. 
#восстановлена кнопка удалить
# добавлена возможность поворота на 90 градусов зажатием ЛКМ+R 
#ВНИМАНИЕ работает только в английской раскладке и в зависимости от заглавности буквы
# исправлены коллизии 
# добавлена строка колличество
# поправлено расположение видов и добавлены надписи
# изменено расположение видов, теперь они расположены как в сиплане
# контейнер заменен на 40фт 

# надо исправить счетчик осталось, не работет кнопка добавить 