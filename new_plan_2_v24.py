# container_app.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl import Workbook
import re
import os
from PIL import Image, ImageTk

INVALID_CHARS = re.compile(r'[,;*?<>|":\\/]')
NON_ASCII = re.compile(r'[^\x00-\x7F]')

# ------------------------------------------------------------------
# Габариты 40-футового контейнера (мм)
L40, W40, H40 = 32_200, 5_000, 5_000
SCALE = 0.035                       # px / mm
# ------------------------------------------------------------------

def validate_name(name: str) -> str:
    name = name.strip()
    if not name:
        raise ValueError("Название не может быть пустым")
    if INVALID_CHARS.search(name):
        raise ValueError("Название содержит запрещённые символы: , ; * ? < > | \" : \\ /")
    if NON_ASCII.search(name):
        raise ValueError("Название содержит иероглифы или спецсимволы")
    return name

MAX_WEIGHT_KG = 28_000
MAX_VOLUME_M3 = 33.2   # 20-футовый контейнер

# ------------------------------------------------------------------
class ContainerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Размещение груза в контейнере")
        self.drag_instance = None
        self.selected = set()
        self.select_rect = None
        self.select_start = None
        self.drag_mode = None
        self.drag_keys = []
        self.drag_offset = None

        # ---- ключи без двоеточий ----
        self.entries = {}

        # габариты контейнера (мм)
        self.L, self.W, self.H = 32200, 5000, 5000
        self.tw, self.th = int(self.L * SCALE), int(self.W * SCALE)
        self.sw, self.sh = int(self.L * SCALE), int(self.H * SCALE)
        self.fw = int(self.W * SCALE * 2)
        self.fh = int(self.H * SCALE * 2)

        self.cargos = []
        self.counter = 1
        self.colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
                       "#9467bd", "#8c564b", "#e377c2", "#7f7f7f",
                       "#bcbd22", "#17becf"]

        self.build_gui()
        self.root.bind_all("<KeyPress-r>", self.on_rotate_key)
        self.tree.bind("<Double-1>", self.on_cell_double_click)
        self.root.bind_all("<Delete>", self.delete_selected_rects)

        self.group_anchor = (0, 0, 0)

        self.bg_paths = {
            'top':   'resources/top.png',
            'side':  'resources/side.png',
            'front': 'resources/front.png'
        }

    # ------------------------------------------------------------------
    # Редактирование ячейки
    # ------------------------------------------------------------------
    def delete_selected_rects(self, event=None):
        if not self.selected:
            return
        for cid, idx in sorted(self.selected, key=lambda x: x[1], reverse=True):
            cargo = next((c for c in self.cargos if c['id'] == cid), None)
            if not cargo or idx >= len(cargo.get('instances', [])):
                continue
            inst = cargo['instances'][idx]
            for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
                canv.delete(inst['top_id'])
                canv.delete(inst['side_id'])
                canv.delete(inst['front_id'])
            cargo['instances'].pop(idx)
            cargo['placed'] = max(0, cargo['placed'] - 1)
            self.refresh_tree_row(cargo)
        self.selected.clear()
        self.recolor_selected()
        self.update_status()

    def on_cell_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return

        editable_cols = ("name", "qty", "l", "w", "h", "wt")
        col_index = int(col_id.replace("#", "")) - 1
        if col_index >= len(editable_cols):
            return
        col_key = editable_cols[col_index]
        current_text = self.tree.set(row_id, col_id)
        x, y, w, h = self.tree.bbox(row_id, col_id)

        if hasattr(self, 'edit_entry') and self.edit_entry.winfo_exists():
            self.edit_entry.destroy()

        self.edit_entry = tk.Entry(self.tree)
        self.edit_entry.place(x=x, y=y, width=w, height=h)
        self.edit_entry.insert(0, current_text)
        self.edit_entry.select_range(0, "end")
        self.edit_entry.focus()

        self._edit_row_id = row_id
        self._edit_col_key = col_key
        self._edit_col_id = col_id

        self.edit_entry.bind("<Return>", self.save_cell_edit)
        self.edit_entry.bind("<FocusOut>", self.save_cell_edit)
        self.edit_entry.bind("<Escape>", self.cancel_cell_edit)

    def save_cell_edit(self, event=None):
        new_val = self.edit_entry.get().strip()
        try:
            col = self._edit_col_key
            cargo = next(c for c in self.cargos if str(c['id']) == self._edit_row_id)
            if col == "name":
                new_val = validate_name(new_val)
                cargo['name'] = new_val
            else:
                new_val = int(new_val)
                if new_val < 0:
                    raise ValueError("Отрицательное значение")
                if col == "qty":
                    cargo['qty'] = new_val
                elif col == "l":
                    cargo['length'] = new_val
                elif col == "w":
                    cargo['width'] = new_val
                elif col == "h":
                    cargo['height'] = new_val
                elif col == "wt":
                    cargo['weight'] = new_val

            self.tree.set(self._edit_row_id, self._edit_col_id, new_val)
            if col in ("l", "w", "h"):
                for inst in cargo.get('instances', []):
                    inst['length'] = cargo['length']
                    inst['width']  = cargo['width']
                    inst['height'] = cargo['height']
                    self.move_instance_to(inst, inst['x'], inst['y'], inst['z'])
            self.update_status()
            self.refresh_tree_row(cargo)
        except ValueError as e:
            messagebox.showerror("Ошибка", str(e))
        finally:
            self.cancel_cell_edit()

    def cancel_cell_edit(self, event=None):
        if hasattr(self, 'edit_entry') and self.edit_entry.winfo_exists():
            self.edit_entry.destroy()
        self.__dict__.pop('edit_entry', None)

    # ------------------------------------------------------------------
    # Размещение партии
    # ------------------------------------------------------------------
    def place_batch(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз в таблице")
            return
        cid = int(sel[0])
        cargo = next(c for c in self.cargos if c['id'] == cid)
        qty_left = cargo['qty'] - cargo['placed']
        if qty_left <= 0:
            messagebox.showinfo("Инфо", "Вся партия уже размещена")
            return

        l, w, h = cargo['length'], cargo['width'], cargo['height']
        gap = 10

        placed_now = 0
        x = y = z = 0

        while placed_now < qty_left:
            end_x = x + l
            end_y = y + w
            end_z = z + h
            if end_y > self.W:
                y = 0
                x += l + gap
                continue
            if end_x > self.L:
                x = y = 0
                z += h + gap
                continue
            if end_z > self.H:
                break
            if self.collision_3d(x, y, z, l, w, h):
                y += w + gap
                continue

            inst = {
                'x': x, 'y': y, 'z': z,
                'length': l, 'width': w, 'height': h,
                'name': cargo['name'], 'weight': cargo['weight']
            }

            inst['top_id'] = self.top_canvas.create_rectangle(
                10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
                10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE,
                fill="#0078D4", outline="black",
                tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

            inst['side_id'] = self.side_canvas.create_rectangle(
                10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
                10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE,
                fill="#0078D4", outline="black",
                tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

            inst['front_id'] = self.front_canvas.create_rectangle(
                10 + y * SCALE * 2, 10 + (self.H - h - z) * SCALE * 2,
                10 + (y + w) * SCALE * 2, 10 + (self.H - z) * SCALE * 2,
                fill="#0078D4", outline="black",
                tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

            cargo.setdefault('instances', []).append(inst)
            cargo['placed'] += 1
            placed_now += 1
            y += w + gap

        if placed_now == 0:
            messagebox.showwarning("Не влезло", "Ни одной коробки не удалось разместить")
        else:
            self.refresh_tree_row(cargo)
            self.update_status()

    # ------------------------------------------------------------------
    def build_gui(self):
        left = tk.Frame(self.root, width=300)
        left.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)
        left.pack_propagate(False)

        frm = tk.LabelFrame(left, text="Параметры груза")
        frm.pack(fill=tk.X, pady=(0, 10))

        cols = tk.Frame(frm)
        cols.pack(pady=5)

        col1 = tk.Frame(cols)
        col1.pack(side=tk.LEFT, padx=5)
        labels1 = ("Название", "Количество", "Вес (кг)")
        for txt in labels1:
            tk.Label(col1, text=txt+":", width=12, anchor='w').pack()
            e = tk.Entry(col1, width=15)
            e.pack()
            self.entries[txt] = e

        col2 = tk.Frame(cols)
        col2.pack(side=tk.LEFT, padx=5)
        labels2 = ("Длина (мм)", "Ширина (мм)", "Высота (мм)")
        for txt in labels2:
            tk.Label(col2, text=txt+":", width=12, anchor='w').pack()
            e = tk.Entry(col2, width=15)
            e.pack()
            self.entries[txt] = e

        btns = tk.Frame(left)
        btns.pack(fill=tk.X, pady=5)

        row1 = tk.Frame(btns)
        row1.pack()
        tk.Button(row1, text="Добавить",  command=self.add_to_table,    width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row1, text="Экспорт",   command=self.export_xlsx,     width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row1, text="Импорт",    command=self.import_xlsx,     width=12).pack(side=tk.LEFT, padx=3)

        row2 = tk.Frame(btns)
        row2.pack()
        tk.Button(row2, text="Разместить", command=self.place_selected, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row2, text="Удалить",    command=self.delete_selected, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(row2, text="Очистить",   command=self.clear_canvas,    width=12).pack(side=tk.LEFT, padx=3)

        row3 = tk.Frame(btns)
        row3.pack()
        tk.Button(row3, text="Разместить партию", command=self.place_batch, width=36).pack(pady=3)

        self.status_lbl = tk.Label(left, text="Вес: 0 / 28 000 кг   Объём: 0 / 33 м³")
        self.status_lbl.pack(pady=5)

        self.tree = ttk.Treeview(left,
                                 columns=("name", "qty", "l", "w", "h", "wt", "placed", "left"),
                                 show="headings", height=10)
        self.tree.heading("name", text="Название")
        self.tree.heading("qty", text="Кол-во")
        self.tree.heading("l", text="Длина")
        self.tree.heading("w", text="Ширина")
        self.tree.heading("h", text="Высота")
        self.tree.heading("wt", text="Вес")
        self.tree.heading("placed", text="Размещено")
        self.tree.heading("left", text="Осталось")
        for col in ("name", "qty", "l", "w", "h", "wt", "placed", "left"):
            self.tree.column(col, width=80, anchor="center")
        self.tree.pack(fill=tk.BOTH, expand=True)

        right = tk.Frame(self.root)
        right.pack(side=tk.RIGHT, padx=5, pady=5, fill=tk.BOTH, expand=True)

        side_box = tk.Frame(right)
        side_box.pack(anchor='w')
        self.side_canvas = tk.Canvas(side_box,
                                     width=self.sw + 20,
                                     height=self.sh + 20,
                                     bg="gray95")
        self.side_canvas.pack()

        top_box = tk.Frame(right)
        top_box.pack(anchor='w')
        self.top_canvas = tk.Canvas(top_box,
                                    width=self.tw + 20,
                                    height=self.th + 20,
                                    bg="gray95")
        self.top_canvas.pack()

        front_box = tk.Frame(right)
        front_box.pack(anchor='w')
        self.front_canvas = tk.Canvas(front_box,
                                      width=self.fw + 20,
                                      height=self.fh + 20,
                                      bg="gray95")
        self.front_canvas.pack()

        self.top_tooltip  = Tooltip(self.top_canvas, self)
        self.side_tooltip = Tooltip(self.side_canvas, self)
        self.front_tooltip = Tooltip(self.front_canvas, self)

        for canv, tag in ((self.top_canvas,  'top'),
                          (self.side_canvas, 'side'),
                          (self.front_canvas,'front')):
            canv.bind("<Button-1>", lambda e, t=tag: self.start_drag(e, t))
            canv.bind("<B1-Motion>", lambda e, t=tag: self.do_drag(e, t))
            canv.bind("<ButtonRelease-1>", lambda e, t=tag: self.end_drag(e, t))
            canv.bind("<Button-3>", lambda e, t=tag: self.select_start_rect(e, t))
            canv.bind("<B3-Motion>", lambda e, t=tag: self.select_drag_rect(e, t))
            canv.bind("<ButtonRelease-3>", lambda e, t=tag: self.select_end_rect(e, t))

        self.root.bind_all("<Escape>", self.clear_selection)

        for c in (self.top_canvas, self.side_canvas, self.front_canvas):
            c.bind("<Double-1>", self.remove_from_canvas)

        self.draw_containers()

    # ------------------------------------------------------------------
    def draw_containers(self):
        tw = max(1, int(self.L * SCALE))
        th = max(1, int(self.W * SCALE))
        sw = max(1, int(self.L * SCALE))
        sh = max(1, int(self.H * SCALE))
        fw = max(1, int(self.W * SCALE * 2))
        fh = max(1, int(self.H * SCALE * 2))

        self.tw, self.th = tw, th
        self.sw, self.sh = sw, sh
        self.fw, self.fh = fw, fh

        base_dir = os.path.dirname(os.path.abspath(__file__))
        bg_dir = os.path.join(base_dir, "resources")
        bg_files = {
            'top':   os.path.join(bg_dir, "top.png"),
            'side':  os.path.join(bg_dir, "side.png"),
            'front': os.path.join(bg_dir, "front.png")
        }

        try:
            self.bg_img = {}
            for key, path in bg_files.items():
                if not os.path.exists(path):
                    raise FileNotFoundError(path)
                img = Image.open(path).convert("RGBA")
                size = {'top': (tw, th),
                        'side': (sw, sh),
                        'front': (fw, fh)}[key]
                self.bg_img[key] = ImageTk.PhotoImage(img.resize(size, Image.LANCZOS))

            self.top_canvas.create_image(10, 10, anchor='nw', image=self.bg_img['top'])
            self.side_canvas.create_image(10, 10, anchor='nw', image=self.bg_img['side'])
            self.front_canvas.create_image(10, 10, anchor='nw', image=self.bg_img['front'])
        except FileNotFoundError as e:
            messagebox.showerror("Файл не найден", f"Не удалось загрузить фон: {e}")

        for canv, w, h in ((self.top_canvas, tw, th),
                           (self.side_canvas, sw, sh),
                           (self.front_canvas, fw, fh)):
            canv.create_rectangle(10, 10, 10 + w, 10 + h, width=2, outline="black")

    # ------------------------------------------------------------------
    def add_to_table(self):
        try:
            name = validate_name(self.entries["Название"].get())
            qty  = int(self.entries["Количество"].get())
            l = int(self.entries["Длина (мм)"].get())
            w = int(self.entries["Ширина (мм)"].get())
            h = int(self.entries["Высота (мм)"].get())
            wt = int(self.entries["Вес (кг)"].get())

            if qty <= 0 or min(l, w, h, wt) <= 0:
                raise ValueError
            if l > self.L or w > self.W or h > self.H:
                messagebox.showerror("Ошибка", "Габариты превышают контейнер")
                return

            cargo = {
                'id': self.counter,
                'name': name,
                'qty': qty,
                'length': l, 'width': w, 'height': h, 'weight': wt,
                'placed': 0,
                'instances': []
            }
            self.cargos.append(cargo)
            self.tree.insert("", "end", iid=str(self.counter),
                             values=(name, qty, l, w, h, wt, 0, qty))
            self.counter += 1
            self.clear_entries()
            self.entries["Название"].focus_set()
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте ввод")

    def refresh_tree_row(self, cargo):
        left = cargo['qty'] - cargo['placed']
        self.tree.item(str(cargo['id']),
                       values=(cargo['name'], cargo['qty'],
                               cargo['length'], cargo['width'], cargo['height'],
                               cargo['weight'], cargo['placed'], left))

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз для удаления")
            return
        cid = int(sel[0])
        cargo = next((c for c in self.cargos if c['id'] == cid), None)
        if not cargo:
            return
        for inst in cargo.get('instances', []):
            for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
                canv.delete(inst['top_id'])
                canv.delete(inst['side_id'])
                canv.delete(inst['front_id'])
        self.tree.delete(sel[0])
        self.cargos = [c for c in self.cargos if c['id'] != cid]
        self.update_status()

    def place_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз в таблице")
            return
        cid = int(sel[0])
        cargo = next(c for c in self.cargos if c['id'] == cid)
        if cargo['placed'] >= cargo['qty']:
            messagebox.showwarning("Лимит", f"Все {cargo['qty']} ед. уже размещены")
            return

        x = (self.L - cargo['length']) // 2
        y = (self.W - cargo['width']) // 2
        z = self.drop_z(x, y, cargo['length'], cargo['width'])

        inst = {
            'x': x, 'y': y, 'z': z,
            'length': cargo['length'], 'width': cargo['width'], 'height': cargo['height'],
            'name': cargo['name'], 'weight': cargo['weight']
        }

        inst['top_id'] = self.top_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.W - inst['width'] - y) * SCALE,
            10 + (x + inst['length']) * SCALE, 10 + (self.W - y) * SCALE,
            fill="#0078D4", outline="black",
            tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

        inst['side_id'] = self.side_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.H - inst['height'] - z) * SCALE,
            10 + (x + inst['length']) * SCALE, 10 + (self.H - z) * SCALE,
            fill="#0078D4", outline="black",
            tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

        inst['front_id'] = self.front_canvas.create_rectangle(
            10 + y * SCALE * 2, 10 + (self.H - inst['height'] - z) * SCALE * 2,
            10 + (y + inst['width']) * SCALE * 2, 10 + (self.H - z) * SCALE * 2,
            fill="#0078D4", outline="black",
            tags=("cargo", f"{cargo['id']}_{len(cargo.get('instances', []))}"))

        cargo.setdefault('instances', []).append(inst)
        cargo['placed'] += 1
        self.refresh_tree_row(cargo)
        self.update_status()

    def remove_from_canvas(self, evt):
        canvas = evt.widget
        items = canvas.find_withtag("current")
        if not items:
            return
        clicked_id = items[0]
        cargo = inst = None
        for c in self.cargos:
            for ins in c.get('instances', []):
                if clicked_id in (ins.get('top_id'), ins.get('side_id'), ins.get('front_id')):
                    cargo = c
                    inst = ins
                    break
            if cargo:
                break
        if not inst:
            return
        for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
            canv.delete(inst.get('top_id', 0))
            canv.delete(inst.get('side_id', 0))
            canv.delete(inst.get('front_id', 0))
        cargo['instances'].remove(inst)
        cargo['placed'] = max(0, cargo['placed'] - 1)
        self.refresh_tree_row(cargo)
        self.update_status()

    def collision_3d(self, x, y, z, l, w, h, exclude_instance=None):
        for c in self.cargos:
            for inst in c.get('instances', []):
                if inst is exclude_instance:
                    continue
                if (x < inst['x'] + inst['length'] and x + l > inst['x'] and
                    y < inst['y'] + inst['width'] and y + w > inst['y'] and
                    z < inst['z'] + inst['height'] and z + h > inst['z']):
                    return True
        return False

    def export_xlsx(self):
        if not self.cargos:
            messagebox.showwarning("Нет данных", "Нечего экспортировать")
            return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Грузы"
        ws.append(["№", "Название", "Количество", "Длина (мм)", "Ширина (мм)", "Высота (мм)", "Вес (кг)"])
        for c in self.cargos:
            ws.append([c['id'], c['name'], c['qty'],
                       c['length'], c['width'], c['height'], c['weight']])
        wb.save(f)
        messagebox.showinfo("Готово", f"Сохранено в {f}")

    def import_xlsx(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 7 or not all(row[:7]):
                    continue
                _, name, qty, l, w, h, wt = row[:7]
                cargo = {
                    'id': self.counter,
                    'name': str(name),
                    'qty': int(qty),
                    'length': int(l), 'width': int(w), 'height': int(h),
                    'weight': int(wt),
                    'placed': 0,
                    'instances': []
                }
                self.cargos.append(cargo)
                self.tree.insert("", "end", iid=str(self.counter),
                                 values=(name, qty, l, w, h, wt, 0, qty))
                self.counter += 1
        except Exception as e:
            messagebox.showerror("Ошибка импорта", str(e))

    def start_drag(self, evt, mode):
        canvas = getattr(self, f"{mode}_canvas")
        items = canvas.find_withtag("current")
        if not items:
            self.drag_keys = []
            return
        clicked_id = items[0]
        if self.selected:
            self.drag_keys = list(self.selected)
        else:
            for c in self.cargos:
                for idx, inst in enumerate(c.get('instances', [])):
                    if inst.get(f"{mode}_id") == clicked_id:
                        self.drag_keys = [(c['id'], idx)]
                        break
                else:
                    continue
                break
            else:
                self.drag_keys = []
                return
        self.drag_start_px = (evt.x, evt.y)
        self.drag_mode = mode

    def do_drag(self, evt, mode):
        if not hasattr(self, 'drag_keys') or not self.drag_keys:
            return
        dx_px = evt.x - self.drag_start_px[0]
        dy_px = evt.y - self.drag_start_px[1]
        if mode == 'front':
            dx_mm = dx_px / (SCALE * 2)
            dy_mm = dy_px / (SCALE * 2)
        else:
            dx_mm = dx_px / SCALE
            dy_mm = dy_px / SCALE

        # 1. Собираем прямоугольники группы и список исключений
        group_rects = []
        exclude_list = [(cid, idx) for cid, idx in self.drag_keys]
        for cid, idx in self.drag_keys:
            c = next(c for c in self.cargos if c['id'] == cid)
            inst = c['instances'][idx]
            if self.drag_mode == 'top':
                gx = max(0, min(self.L - inst['length'], inst['x'] + dx_mm))
                gy = max(0, min(self.W - inst['width'],  inst['y'] - dy_mm))
            elif self.drag_mode == 'side':
                gx = max(0, min(self.L - inst['length'], inst['x'] + dx_mm))
                gy = inst['y']
            else:  # front
                gx = inst['x']
                gy = max(0, min(self.W - inst['width'],  inst['y'] + dx_mm))
            group_rects.append((gx, gy, inst['length'], inst['width']))

        # 2. Один общий Z для всей группы
        common_z = 0
        for gx, gy, gl, gw in group_rects:
            common_z = max(common_z,
                           self.drop_z(gx, gy, gl, gw, exclude_list=exclude_list))

        # 3. Применяем ко всем один и тот же Z
        for cid, idx in self.drag_keys:
            c = next(c for c in self.cargos if c['id'] == cid)
            inst = c['instances'][idx]
            if self.drag_mode == 'top':
                new_x = max(0, min(self.L - inst['length'], inst['x'] + dx_mm))
                new_y = max(0, min(self.W - inst['width'],  inst['y'] - dy_mm))
            elif self.drag_mode == 'side':
                new_x = max(0, min(self.L - inst['length'], inst['x'] + dx_mm))
                new_y = inst['y']
            else:   # front
                new_x = inst['x']
                new_y = max(0, min(self.W - inst['width'],  inst['y'] + dx_mm))
            self.move_instance_to(inst, new_x, new_y, common_z)

        self.drag_start_px = (evt.x, evt.y)

    def end_drag(self, evt, mode):
        self.drag_keys = []

    def move_instance_to(self, inst, x, y, z):
        inst['x'], inst['y'], inst['z'] = x, y, z
        l, w, h = inst['length'], inst['width'], inst['height']
        self.top_canvas.coords(inst['top_id'],
            10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE)
        self.side_canvas.coords(inst['side_id'],
            10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE)
        self.front_canvas.coords(inst['front_id'],
            10 + y * SCALE * 2, 10 + (self.H - h - z) * SCALE * 2,
            10 + (y + w) * SCALE * 2, 10 + (self.H - z) * SCALE * 2)

    def update_status(self):
        placed = [c for c in self.cargos if c.get('placed', 0) > 0]
        total_w = sum(c['weight'] * len(c.get('instances', [])) for c in placed)
        total_v = sum(inst['length'] * inst['width'] * inst['height']
                      for c in placed for inst in c.get('instances', [])) / 1e9
        self.status_lbl.config(text=f"Вес: {total_w} / {MAX_WEIGHT_KG} кг   "
                                    f"Объём: {total_v:.2f} / {MAX_VOLUME_M3} м³")

    def on_rotate_key(self, evt):
        if len(self.selected) != 1:
            return
        cid, idx = next(iter(self.selected))
        c = next(c for c in self.cargos if c['id'] == cid)
        if not c.get('instances') or idx >= len(c['instances']):
            return
        inst = c['instances'][idx]
        items = self.top_canvas.find_withtag("current")
        if not items or items[0] != inst.get('top_id'):
            return

        new_l = inst['width']
        new_w = inst['length']
        h = inst['height']
        if (inst['x'] + new_l > self.L) or (inst['y'] + new_w > self.W):
            messagebox.showerror("Ошибка", "Поворот выходит за границы")
            return
        if self.collision_3d(inst['x'], inst['y'], inst['z'], new_l, new_w, h, exclude_instance=inst):
            messagebox.showerror("Ошибка", "Поворот приведёт к пересечению")
            return
        new_z = self.drop_z(inst['x'], inst['y'], new_l, new_w, exclude_instance=inst)
        inst['length'], inst['width'] = new_l, new_w
        self.move_instance_to(inst, inst['x'], inst['y'], new_z)
        self.recolor_selected()

    def drop_z(self, x, y, l, w, exclude_list=None):
        """exclude_list = список кортежей (cid, idx) исключений."""
        if exclude_list is None:
            exclude_list = []
        max_z = 0
        for c in self.cargos:
            for idx, inst in enumerate(c.get('instances', [])):
                if (c['id'], idx) in exclude_list:
                    continue
                if (x < inst['x'] + inst['length'] and x + l > inst['x'] and
                    y < inst['y'] + inst['width'] and y + w > inst['y']):
                    max_z = max(max_z, inst['z'] + inst['height'])
        return max_z

    def clear_canvas(self):
        for canv in (self.top_canvas, self.side_canvas, self.front_canvas):
            canv.delete("cargo")
        for c in self.cargos:
            c['placed'] = 0
            c['instances'] = []
            self.refresh_tree_row(c)
        self.update_status()

    def clear_entries(self):
        for e in self.entries.values():
            e.delete(0, tk.END)

    def select_start_rect(self, evt, mode):
        if not (evt.state & 0x0001):
            self.clear_selection()
        self.select_mode = mode
        self.select_start = (evt.x, evt.y)
        canv = getattr(self, f"{mode}_canvas")
        if self.select_rect and canv.winfo_exists():
            canv.delete(self.select_rect)
        self.select_rect = None

    def select_drag_rect(self, evt, mode):
        canv = getattr(self, f"{mode}_canvas")
        if self.select_rect and canv.winfo_exists():
            canv.delete(self.select_rect)
        x0, y0 = self.select_start
        self.select_rect = canv.create_rectangle(
            x0, y0, evt.x, evt.y, outline="green", width=2, dash=(4, 2))

    def select_end_rect(self, evt, mode):
        if not self.select_start:
            return
        canv = getattr(self, f"{mode}_canvas")
        if self.select_rect and canv.winfo_exists():
            canv.delete(self.select_rect)
        self.select_rect = None
        x0, y0 = self.select_start
        x1, y1 = evt.x, evt.y
        x0, x1 = min(x0, x1), max(x0, x1)
        y0, y1 = min(y0, y1), max(y0, y1)

        new_selection = set()
        for c in self.cargos:
            for idx, inst in enumerate(c.get('instances', [])):
                px0, py0, px1, py1 = self._inst_bbox(inst, mode)
                if not (x1 < px0 or x0 > px1 or y1 < py0 or y0 > py1):
                    new_selection.add((c['id'], idx))
        self.selected = new_selection
        self.recolor_selected()
        self.select_start = None

    def clear_selection(self, evt=None):
        for cid, idx in list(self.selected):
            cargo = next(c for c in self.cargos if c['id'] == cid)
            if idx < len(cargo.get('instances', [])):
                inst = cargo['instances'][idx]
                self.top_canvas.itemconfig(inst['top_id'], fill="#0078D4")
                self.side_canvas.itemconfig(inst['side_id'], fill="#0078D4")
                self.front_canvas.itemconfig(inst['front_id'], fill="#0078D4")
        self.selected.clear()

    def _inst_bbox(self, inst, mode):
        x, y, z, l, w, h = inst['x'], inst['y'], inst['z'], inst['length'], inst['width'], inst['height']
        if mode == 'top':
            px0 = 10 + x * SCALE
            py0 = 10 + (self.W - w - y) * SCALE
            px1 = 10 + (x + l) * SCALE
            py1 = 10 + (self.W - y) * SCALE
        elif mode == 'side':
            px0 = 10 + x * SCALE
            py0 = 10 + (self.H - h - z) * SCALE
            px1 = 10 + (x + l) * SCALE
            py1 = 10 + (self.H - z) * SCALE
        elif mode == 'front':
            px0 = 10 + y * SCALE * 2
            py0 = 10 + (self.H - h - z) * SCALE * 2
            px1 = 10 + (y + w) * SCALE * 2
            py1 = 10 + (self.H - z) * SCALE * 2
        return px0, py0, px1, py1

    def recolor_selected(self):
        for c in self.cargos:
            for idx, inst in enumerate(c.get('instances', [])):
                color = "#00B050" if (c['id'], idx) in self.selected else "#0078D4"
                self.top_canvas.itemconfig(inst['top_id'], fill=color)
                self.side_canvas.itemconfig(inst['side_id'], fill=color)
                self.front_canvas.itemconfig(inst['front_id'], fill=color)

# ------------------------------------------------------------------
class Tooltip:
    def __init__(self, canvas, app):
        self.canvas = canvas
        self.app = app
        self.tooltip = None
        canvas.bind("<Motion>", self.on_motion)
        canvas.bind("<Leave>", self.hide)

    def on_motion(self, event):
        if not self.canvas.winfo_exists():
            return
        items = event.widget.find_withtag("current")
        if not items:
            self.hide()
            return
        clicked_id = items[0]
        cargo = inst = None
        for c in self.app.cargos:
            for ins in c.get('instances', []):
                if clicked_id in (ins.get('top_id'), ins.get('side_id'), ins.get('front_id')):
                    cargo = c
                    inst = ins
                    break
            if cargo:
                break
        if cargo:
            text = f"{cargo['name']}  {inst['weight']} кг"
            self.show(text, event.x_root, event.y_root)
        else:
            self.hide()

    def show(self, text, x, y):
        if self.tooltip and self.tooltip.winfo_exists():
            self.label.config(text=text)
        else:
            self.tooltip = tk.Toplevel(self.canvas)
            self.tooltip.wm_overrideredirect(True)
            self.label = tk.Label(self.tooltip, text=text,
                                  bg="yellow", relief="solid", bd=1)
            self.label.pack()
        self.tooltip.wm_geometry(f"+{x + 10}+{y + 10}")

    def hide(self, *args):
        if self.tooltip and self.tooltip.winfo_exists():
            self.tooltip.destroy()
        self.tooltip = None

# ------------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    ContainerApp(root)
    root.mainloop()


    #крайняя рабочая версия приложения. Добавлены типсы. Добавлено тримирование названий. 
#восстановлена кнопка удалить
# добавлена возможность поворота на 90 градусов зажатием ЛКМ+R 
#ВНИМАНИЕ работает только в английской раскладке и в зависимости от заглавности буквы
# исправлены коллизии 
# добавлена строка колличество
# поправлено расположение видов и добавлены надписи
# изменено расположение видов, теперь они расположены как в сиплане
# контейнер заменен на 40фт 

# надо исправить счетчик осталось, не работет кнопка добавить 
# программа полностью пересобрана kimi 
# добавлена возможность редактирования груза в таблице
#  Как использовать
# Запустите приложение.
# Дважды кликните на любую ячейку строки в таблице.
# Откроется модальное окно с предзаполненными полями.
# Измените значения и нажмите «Сохранить» — все экземпляры этого груза будут удалены и можно будет заново разместить его с новыми размерами.
# добавлена перерисовка после изменения размеров 
# добавлена возможность редактировать груз напрямую в таблице
# добавлена возможность добавлять партию груза
# 26_07_2025
# программа пересобрана с помощью kimi
# исправлено перемещение партии груза