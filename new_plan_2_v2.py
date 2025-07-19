import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl import Workbook
import os

class ContainerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Размещение груза в контейнере")
        
        # Внутренние размеры стандартного 20-футового контейнера в миллиметрах 
        self.container_length_mm = 5905  # Длина
        self.container_width_mm = 2350   # Ширина
        self.container_height_mm = 2381  # Высота
        
        # Масштаб для отображения (1 мм = 0.1 пикселя)
        self.scale = 0.1
        
        # Размеры контейнера на холстах
        self.top_view_width = int(self.container_length_mm * self.scale)
        self.top_view_height = int(self.container_width_mm * self.scale)
        self.side_view_width = int(self.container_length_mm * self.scale)
        self.side_view_height = int(self.container_height_mm * self.scale)
        self.front_view_width = int(self.container_width_mm * self.scale)
        self.front_view_height = int(self.container_height_mm * self.scale)
        
        # Создаем основную рамку для ввода параметров и таблицы
        self.left_frame = tk.Frame(root)
        self.left_frame.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)
        
        # Фрейм для ввода параметров груза
        self.input_frame = tk.Frame(self.left_frame)
        self.input_frame.pack(fill=tk.X)
        
        # Поля ввода параметров груза
        tk.Label(self.input_frame, text="Название груза:").pack()
        self.name_entry = tk.Entry(self.input_frame)
        self.name_entry.pack()
        
        tk.Label(self.input_frame, text="Длина груза (мм):").pack()
        self.length_entry = tk.Entry(self.input_frame)
        self.length_entry.pack()
        
        tk.Label(self.input_frame, text="Ширина груза (мм):").pack()
        self.width_entry = tk.Entry(self.input_frame)
        self.width_entry.pack()
        
        tk.Label(self.input_frame, text="Высота груза (мм):").pack()
        self.height_entry = tk.Entry(self.input_frame)
        self.height_entry.pack()
        
        tk.Label(self.input_frame, text="Вес груза (кг):").pack()
        self.weight_entry = tk.Entry(self.input_frame)
        self.weight_entry.pack()
        
        # Фрейм для кнопок добавления, экспорта и импорта а так же кнопки удаления
        self.button_frame = tk.Frame(self.input_frame)
        self.button_frame.pack(pady=10)
        
        self.add_button = tk.Button(self.button_frame, text="Добавить груз", command=self.add_cargo)
        self.add_button.pack(side=tk.LEFT, padx=5)

        self.delete_button = tk.Button(self.button_frame, text="Удалить выбранный",command=self.delete_selected_cargo)
        self.delete_button.pack(side=tk.LEFT, padx=5)
        
        self.export_button = tk.Button(self.button_frame, text="Экспорт XLSX", command=self.export_to_excel)
        self.export_button.pack(side=tk.LEFT, padx=5)
        
        self.import_button = tk.Button(self.button_frame, text="Импорт XLSX", command=self.import_from_excel)
        self.import_button.pack(side=tk.LEFT, padx=5)
        
        # Таблица грузов
        self.table_frame = tk.Frame(self.left_frame)
        self.table_frame.pack(fill=tk.BOTH, expand=True)
        
        self.tree = ttk.Treeview(self.table_frame, columns=("name", "length", "width", "height", "weight"), 
                                show="headings", height=10)
        
        self.tree.heading("#0", text="№")
        self.tree.heading("name", text="Название")
        self.tree.heading("length", text="Длина (мм)")
        self.tree.heading("width", text="Ширина (мм)")
        self.tree.heading("height", text="Высота (мм)")
        self.tree.heading("weight", text="Вес (кг)")
        
        self.tree.column("#0", width=40)
        self.tree.column("name", width=100)
        self.tree.column("length", width=80)
        self.tree.column("width", width=80)
        self.tree.column("height", width=80)
        self.tree.column("weight", width=80)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Создаем рамку для видов контейнера
        self.view_frame = tk.Frame(root)
        self.view_frame.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        # Верхний фрейм для видов спереди и сбоку
        self.top_views_frame = tk.Frame(self.view_frame)
        self.top_views_frame.pack(fill=tk.X)
        
        # Вид спереди
        self.front_view_frame = tk.Frame(self.top_views_frame)
        self.front_view_frame.pack(side=tk.LEFT, padx=5)
        self.front_view_label = tk.Label(self.front_view_frame, text="Вид спереди")
        self.front_view_label.pack()
        self.front_canvas = tk.Canvas(self.front_view_frame, width=self.front_view_width + 20, 
                                     height=self.front_view_height + 20, bg="lightgray")
        self.front_canvas.pack()
        
        # Вид сбоку
        self.side_view_frame = tk.Frame(self.top_views_frame)
        self.side_view_frame.pack(side=tk.LEFT, padx=5)
        self.side_view_label = tk.Label(self.side_view_frame, text="Вид сбоку")
        self.side_view_label.pack()
        self.side_canvas = tk.Canvas(self.side_view_frame, width=self.side_view_width + 20, 
                                    height=self.side_view_height + 20, bg="lightgray")
        self.side_canvas.pack()
        
        # Вид сверху
        self.top_view_frame = tk.Frame(self.view_frame)
        self.top_view_frame.pack(pady=10)
        self.top_view_label = tk.Label(self.top_view_frame, text="Вид сверху")
        self.top_view_label.pack()
        self.top_canvas = tk.Canvas(self.top_view_frame, width=self.top_view_width + 20, 
                                   height=self.top_view_height + 20, bg="lightgray")
        self.top_canvas.pack()
        
        # Рисуем контейнер (вид спереди)
        self.front_container = self.front_canvas.create_rectangle(
            10, 10, self.front_view_width + 10, self.front_view_height + 10,
            outline="black", width=2, fill="white"
        )
        
        # Рисуем контейнер (вид сбоку)
        self.side_container = self.side_canvas.create_rectangle(
            10, 10, self.side_view_width + 10, self.side_view_height + 10,
            outline="black", width=2, fill="white"
        )
        
        # Рисуем контейнер (вид сверху)
        self.top_container = self.top_canvas.create_rectangle(
            10, 10, self.top_view_width + 10, self.top_view_height + 10,
            outline="black", width=2, fill="white"
        )
        
        # Переменные для перетаскивания
        self.selected_cargo = None
        self.offset_x = 0
        self.offset_y = 0
        
        # Привязываем события мыши только к виду сверху
        self.top_canvas.bind("<Button-1>", self.on_click)
        self.top_canvas.bind("<B1-Motion>", self.on_drag)
        self.top_canvas.bind("<ButtonRelease-1>", self.on_release)
        for canvas in (self.top_canvas, self.side_canvas, self.front_canvas):
            canvas.bind("<Double-1>", self.erase_cargo_on_canvas)
        
        # Список для хранения всех грузов
        self.cargos = []
        self.cargo_counter = 1
        
        # Доступные цвета для грузов
        self.colors = ["blue", "green", "red", "orange", "purple", "cyan", "magenta", "yellow"]
    
    def get_random_color(self):
        return random.choice(self.colors)
    
    def clear_all_cargos(self):
        """Удаляет все грузы из интерфейса и очищает список"""
        for cargo in self.cargos:
            self.top_canvas.delete(cargo['top_id'])
            self.side_canvas.delete(cargo['side_id'])
            self.front_canvas.delete(cargo['front_id'])
        self.cargos = []
        self.cargo_counter = 1
        for item in self.tree.get_children():
            self.tree.delete(item)
    
    def add_cargo(self):
        try:
            name = self.name_entry.get()
            if not name:
                messagebox.showerror("Ошибка", "Введите название груза")
                return
                
            length_mm = int(self.length_entry.get())
            width_mm = int(self.width_entry.get())
            height_mm = int(self.height_entry.get())
            weight = int(self.weight_entry.get())
            
            if length_mm <= 0 or width_mm <= 0 or height_mm <= 0 or weight <= 0:
                messagebox.showerror("Ошибка", "Все параметры должны быть положительными числами")
                return
                
            # Преобразуем мм в пиксели
            length = int(length_mm * self.scale)
            width = int(width_mm * self.scale)
            height_px = int(height_mm * self.scale)
            
            # Проверяем, помещается ли груз в контейнер
            if (length > self.container_length_mm or 
                width > self.container_width_mm or 
                height_mm > self.container_height_mm):
                messagebox.showerror("Ошибка", "Груз не помещается в контейнер")
                return
                
            # Создаем груз в центре контейнера (вид сверху)
            x1 = 10 + (self.top_view_width - length) // 2
            y1 = 10 + (self.top_view_height - width) // 2
            x2 = x1 + length
            y2 = y1 + width
            
            color = self.get_random_color()
            
            # Создаем груз на виде сверху
            top_cargo = self.top_canvas.create_rectangle(x1, y1, x2, y2, fill=color, tags="cargo")
            
            # Создаем груз на виде сбоку (высота отображается)
            side_x1 = x1  # Та же координата X, что и на виде сверху
            side_y1 = 10 + self.side_view_height - height_px  # Высота от нижней границы
            side_x2 = x2
            side_y2 = 10 + self.side_view_height  # Нижняя граница контейнера
            
            side_cargo = self.side_canvas.create_rectangle(side_x1, side_y1, side_x2, side_y2, fill=color, tags="cargo")
            
            # Создаем груз на виде спереди (ширина и высота)
            front_x1 = 10 + (self.front_view_width - width) // 2  # Центрируем по ширине
            front_y1 = 10 + self.front_view_height - height_px  # Высота от нижней границы
            front_x2 = front_x1 + width
            front_y2 = 10 + self.front_view_height  # Нижняя граница контейнера
            
            front_cargo = self.front_canvas.create_rectangle(front_x1, front_y1, front_x2, front_y2, fill=color, tags="cargo")
            
            # Сохраняем информацию о грузе
            cargo_data = {
                'number': self.cargo_counter,
                'name': name,
                'top_id': top_cargo,
                'side_id': side_cargo,
                'front_id': front_cargo,
                'color': color,
                'length_mm': length_mm,
                'width_mm': width_mm,
                'height_mm': height_mm,
                'weight': weight,
                'length': length,
                'width': width,
                'height_px': height_px,
                'x1': x1,
                'y1': y1
            }
            
            self.cargos.append(cargo_data)
            
            # Добавляем груз в таблицу
            self.tree.insert("", "end", text=str(self.cargo_counter), 
                            values=(name, length_mm, width_mm, height_mm, weight))
            
            self.cargo_counter += 1
            
            # Очищаем поля ввода
            self.name_entry.delete(0, tk.END)
            self.length_entry.delete(0, tk.END)
            self.width_entry.delete(0, tk.END)
            self.height_entry.delete(0, tk.END)
            self.weight_entry.delete(0, tk.END)
            
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")
    
    def delete_selected_cargo(self):  #Метод удаления выбранного груза из таблицы Добавьте в клас
        selection = self.tree.selection()
        if not selection:
            messagebox.showerror("Ошибка", "Выберите груз в таблице для удаления")
            return

        item = selection[0]
        values = self.tree.item(item, "values")
        name = values[0]  # название груза

        # Найдём соответствующий груз в списке cargos
        cargo_to_remove = None
        for cargo in self.cargos:
            if cargo['name'] == name:
                cargo_to_remove = cargo
                break

        if cargo_to_remove:
            # Удаляем с холстов
            self.top_canvas.delete(cargo_to_remove['top_id'])
            self.side_canvas.delete(cargo_to_remove['side_id'])
            self.front_canvas.delete(cargo_to_remove['front_id'])

            # Удаляем из списка и таблицы
            self.cargos.remove(cargo_to_remove)
            self.tree.delete(item)

    def export_to_excel(self):
        """Экспортирует таблицу грузов в файл XLSX"""
        if not self.cargos:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить файл"
        )
        
        if not file_path:
            return  # Пользователь отменил сохранение
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Грузы в контейнере"
            
            # Заголовки столбцов
            headers = ["№", "Название", "Длина (мм)", "Ширина (мм)", "Высота (мм)", "Вес (кг)"]
            ws.append(headers)
            
            # Данные
            for cargo in self.cargos:
                ws.append([
                    cargo['number'],
                    cargo['name'],
                    cargo['length_mm'],
                    cargo['width_mm'],
                    cargo['height_mm'],
                    cargo['weight']
                ])
            
            # Автоматическая ширина столбцов
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            messagebox.showinfo("Успех", f"Данные успешно экспортированы в {file_path}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные: {str(e)}")
    
    def import_from_excel(self):
        """Импортирует таблицу грузов из файла XLSX"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Выберите файл для импорта"
        )
        
        if not file_path:
            return  # Пользователь отменил выбор файла
            
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Очищаем текущие данные
            self.clear_all_cargos()
            
            # Пропускаем заголовок и читаем данные
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 6 and all(row[1:6]):  # Проверяем, что есть данные
                    # Заполняем поля ввода
                    self.name_entry.delete(0, tk.END)
                    self.name_entry.insert(0, str(row[1]))
                    
                    self.length_entry.delete(0, tk.END)
                    self.length_entry.insert(0, str(row[2]))
                    
                    self.width_entry.delete(0, tk.END)
                    self.width_entry.insert(0, str(row[3]))
                    
                    self.height_entry.delete(0, tk.END)
                    self.height_entry.insert(0, str(row[4]))
                    
                    self.weight_entry.delete(0, tk.END)
                    self.weight_entry.insert(0, str(row[5]))
                    
                    # Добавляем груз
                    self.add_cargo()
            
            messagebox.showinfo("Успех", f"Данные успешно импортированы из {file_path}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать данные: {str(e)}")
    
    def on_click(self, event):
        # Находим все грузы в точке клика (вид сверху)
        items = self.top_canvas.find_overlapping(event.x, event.y, event.x, event.y)
        
        if items:
            # Берем верхний груз (последний в списке)
            clicked_id = items[-1]
            # Находим соответствующий груз в нашем списке
            for cargo in self.cargos:
                if cargo['top_id'] == clicked_id:
                    self.selected_cargo = cargo
                    # Запоминаем смещение курсора относительно верхнего левого угла груза
                    cargo_coords = self.top_canvas.coords(self.selected_cargo['top_id'])
                    self.offset_x = event.x - cargo_coords[0]
                    self.offset_y = event.y - cargo_coords[1]
                    break
    
    def on_drag(self, event):
        if self.selected_cargo:
            # Предварительные координаты (вид сверху)
            x1 = event.x - self.offset_x
            y1 = event.y - self.offset_y
            x2 = x1 + self.selected_cargo['length']
            y2 = y1 + self.selected_cargo['width']
            
            # Проверка границ контейнера (вид сверху)
            if x1 < 10:  # Левая граница
                x1 = 10
                x2 = x1 + self.selected_cargo['length']
            if y1 < 10:  # Верхняя граница
                y1 = 10
                y2 = y1 + self.selected_cargo['width']
            if x2 > 10 + self.top_view_width:  # Правая граница
                x2 = 10 + self.top_view_width
                x1 = x2 - self.selected_cargo['length']
            if y2 > 10 + self.top_view_height:  # Нижняя граница
                y2 = 10 + self.top_view_height
                y1 = y2 - self.selected_cargo['width']
            
            # Проверка пересечения с другими грузами (вид сверху)
            for cargo in self.cargos:
                if cargo['top_id'] != self.selected_cargo['top_id']:
                    coords = self.top_canvas.coords(cargo['top_id'])
                    # Если есть пересечение, корректируем позицию
                    if not (x2 < coords[0] or x1 > coords[2] or y2 < coords[1] or y1 > coords[3]):
                        # Определяем направление наименьшего перекрытия
                        overlap_left = x2 - coords[0]
                        overlap_right = coords[2] - x1
                        overlap_top = y2 - coords[1]
                        overlap_bottom = coords[3] - y1
                        
                        min_overlap = min(overlap_left, overlap_right, overlap_top, overlap_bottom)
                        
                        if min_overlap == overlap_left:
                            x1 = coords[0] - self.selected_cargo['length']
                            x2 = coords[0]
                        elif min_overlap == overlap_right:
                            x1 = coords[2]
                            x2 = coords[2] + self.selected_cargo['length']
                        elif min_overlap == overlap_top:
                            y1 = coords[1] - self.selected_cargo['width']
                            y2 = coords[1]
                        elif min_overlap == overlap_bottom:
                            y1 = coords[3]
                            y2 = coords[3] + self.selected_cargo['width']
            
            # Обновляем позицию груза на виде сверху
            self.top_canvas.coords(self.selected_cargo['top_id'], x1, y1, x2, y2)
            
            # Обновляем позицию груза на виде сбоку (только X координата)
            side_y1 = 10 + self.side_view_height - self.selected_cargo['height_px']
            side_y2 = 10 + self.side_view_height
            self.side_canvas.coords(self.selected_cargo['side_id'], x1, side_y1, x2, side_y2)
            
            # Обновляем позицию груза на виде спереди (Y координата - высота остается той же)
            front_width = self.selected_cargo['width']
            front_x1 = y1  # Координата Y на виде сверху становится координатой X на виде спереди
            front_x2 = front_x1 + front_width
            front_y1 = 10 + self.front_view_height - self.selected_cargo['height_px']
            front_y2 = 10 + self.front_view_height
            self.front_canvas.coords(self.selected_cargo['front_id'], front_x1, front_y1, front_x2, front_y2)
            
            # Сохраняем новые координаты
            self.selected_cargo['x1'] = x1
            self.selected_cargo['y1'] = y1
    
    def on_release(self, event):
        self.selected_cargo = None

    def erase_cargo_on_canvas(self, event):
        canvas = event.widget
        items = canvas.find_overlapping(event.x, event.y, event.x, event.y)
        if not items:
            return

        # Найдём cargo по id прямоугольника
        clicked_id = items[-1]
        for cargo in self.cargos:
            if (cargo['top_id'] == clicked_id or
                cargo['side_id'] == clicked_id or
                cargo['front_id'] == clicked_id):

                # Удаляем только прямоугольники
                self.top_canvas.delete(cargo['top_id'])
                self.side_canvas.delete(cargo['side_id'])
                self.front_canvas.delete(cargo['front_id'])

                # Прямоугольники можно заново нарисовать при желании,
                # но строка в таблице остаётся
                break

if __name__ == "__main__":
    root = tk.Tk()
    app = ContainerApp(root)
    root.mainloop()
