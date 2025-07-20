import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl import Workbook
import re
INVALID_CHARS = re.compile(r'[,;*?<>|":\\/]')   # запрещённые символы
NON_ASCII = re.compile(r'[^\x00-\x7F]')          # всё, что вне ASCII (иероглифы, эмодзи)

def validate_name(name: str) -> str:
    """
    Удаляет пробелы по краям, проверяет запрещённые символы.
    Возвращает имя или выбрасывает ValueError.
    """
    name = name.strip()
    if not name:
        raise ValueError("Название не может быть пустым")
    if INVALID_CHARS.search(name):
        raise ValueError("Название содержит запрещённые символы: , ; * ? < > | \" : \\ /")
    if NON_ASCII.search(name):
        raise ValueError("Название содержит иероглифы или спецсимволы")
    return name


MAX_WEIGHT_KG = 28_000
MAX_VOLUME_M3 = 33.2   # 20-футовый контейнер

SCALE = 0.1  # px / mm (фиксированный, без изменений)

# ------------------------------------------------------------------
class ContainerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Размещение груза в контейнере")

        # габариты контейнера (мм)
        self.L, self.W, self.H = 5905, 2350, 2381
        self.tw, self.th = int(self.L * SCALE), int(self.W * SCALE)
        self.sw, self.sh = int(self.L * SCALE), int(self.H * SCALE)
        self.fw, self.fh = int(self.W * SCALE), int(self.H * SCALE)

        self.cargos = []          # все грузы
        self.counter = 1          # для генерации id
        self.colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
                       "#9467bd", "#8c564b", "#e377c2", "#7f7f7f",
                       "#bcbd22", "#17becf"]

        self.build_gui()
        self.draw_containers()
        self.root.bind_all("<KeyPress-r>", self.on_rotate_key)  # <-- сюда

    # ------------------------------------------------------------------
    def build_gui(self):
        # левая панель
        left = tk.Frame(self.root)
        left.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)

        # ввод
        frm = tk.LabelFrame(left, text="Параметры груза")
        frm.pack(fill=tk.X, pady=(0, 10))
        labels = ["Название:", "Длина (мм):", "Ширина (мм):",
                  "Высота (мм):", "Вес (кг):"]
        self.entries = {}
        for lab in labels:
            tk.Label(frm, text=lab).pack()
            e = tk.Entry(frm)
            e.pack()
            self.entries[lab] = e

        # кнопки
        btn_bar = tk.Frame(frm)
        btn_bar.pack(pady=5)
        tk.Button(btn_bar, text="Добавить в таблицу", command=self.add_to_table).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="Разместить", command=self.place_selected).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="Экспорт", command=self.export_xlsx).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="Импорт", command=self.import_xlsx).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="Удалить", command=self.delete_selected).pack(side=tk.LEFT, padx=2)

        # статус
        self.status_lbl = tk.Label(left, text="Вес: 0 / 28 000 кг   Объём: 0 / 33 м³")
        self.status_lbl.pack(pady=5)

        # таблица
        self.tree = ttk.Treeview(left, columns=("name", "l", "w", "h", "wt"),
                                 show="headings", height=10)
        self.tree.heading("name", text="Название")
        self.tree.heading("l", text="Длина (мм)")
        self.tree.heading("w", text="Ширина (мм)")
        self.tree.heading("h", text="Высота (мм)")
        self.tree.heading("wt", text="Вес (кг)")
        for col in ("name", "l", "w", "h", "wt"):
            self.tree.column(col, width=90)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # правая панель — Canvas
        right = tk.Frame(self.root)
        right.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)

        top_frm = tk.Frame(right)
        top_frm.pack(fill=tk.X)

        self.front_canvas = tk.Canvas(top_frm, width=self.fw + 20,
                                      height=self.fh + 20, bg="gray95")
        self.front_canvas.pack(side=tk.LEFT, padx=5)

        self.side_canvas = tk.Canvas(top_frm, width=self.sw + 20,
                                     height=self.sh + 20, bg="gray95")
        self.side_canvas.pack(side=tk.LEFT, padx=5)

        self.top_canvas = tk.Canvas(right, width=self.tw + 20,
                                    height=self.th + 20, bg="gray95")
        self.top_canvas.pack(pady=5)

        # ⬇⬇⬇  ВСТАВЛЯЕМ СЮДА ⬇⬇⬇
        self.top_tooltip  = Tooltip(self.top_canvas, self)
        self.side_tooltip = Tooltip(self.side_canvas, self)
        self.front_tooltip = Tooltip(self.front_canvas, self)
        # ⬆⬆⬆  КОНЕЦ ВСТАВКИ  ⬆⬆⬆

        # bindings
        self.top_canvas.bind("<Button-1>", self.start_drag)
        self.top_canvas.bind("<B1-Motion>", self.do_drag)
        self.top_canvas.bind("<ButtonRelease-1>", self.end_drag)

        # двойной клик по прямоугольнику убирает груз
        for c in (self.top_canvas, self.side_canvas, self.front_canvas):
            c.bind("<Double-1>", self.remove_from_canvas)

    # ------------------------------------------------------------------
    def draw_containers(self):
        for c, w, h in ((self.top_canvas, self.tw, self.th),
                        (self.side_canvas, self.sw, self.sh),
                        (self.front_canvas, self.fw, self.fh)):
            c.create_rectangle(10, 10, 10 + w, 10 + h,
                               width=2, outline="black", fill="white")
        # в конце build_gui
     
    # ------------------------------------------------------------------
    # CRUD
    # ------------------------------------------------------------------
    def add_to_table(self):
        try:
            name = validate_name(self.entries["Название:"].get())
            l = int(self.entries["Длина (мм):"].get())
            w = int(self.entries["Ширина (мм):"].get())
            h = int(self.entries["Высота (мм):"].get())
            wt = int(self.entries["Вес (кг):"].get())

            if not name or min(l, w, h, wt) <= 0:
                raise ValueError
            if l > self.L or w > self.W or h > self.H:
                messagebox.showerror("Ошибка", "Габариты превышают контейнер")
                return

            cargo = {
                'id': self.counter,
                'name': name,
                'length': l, 'width': w, 'height': h, 'weight': wt,
                'placed': False
            }
            self.cargos.append(cargo)
            self.tree.insert("", "end", iid=str(self.counter),
                             values=(name, l, w, h, wt))
            self.counter += 1
            self.clear_entries()
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте ввод")

    # ------------------------------------------------------------------
    # метод удаления груза из таблицы для кнопки удалить
    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз для удаления")
            return

        cid = int(sel[0])
        cargo = next((c for c in self.cargos if c['id'] == cid), None)

        # если груз уже размещён — убираем с холстов
        if cargo and cargo.get('placed'):
            self.top_canvas.delete(cargo.get('top_id', 0))
            self.side_canvas.delete(cargo.get('side_id', 0))
            self.front_canvas.delete(cargo.get('front_id', 0))
            self.update_status()

        # удаляем строку из Treeview
        self.tree.delete(sel[0])
        # удаляем из списка
        self.cargos = [c for c in self.cargos if c['id'] != cid]

    def place_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Ошибка", "Выберите груз в таблице")
            return
        cid = int(sel[0])
        cargo = next(c for c in self.cargos if c['id'] == cid)

        # центр контейнера
        x = (self.L - cargo['length']) // 2
        y = (self.W - cargo['width']) // 2
        z = 0

        # проверка столкновения
        if self.collision_3d(x, y, z,
                             cargo['length'], cargo['width'], cargo['height'],
                             exclude_id=cargo['id']):
            messagebox.showerror("Ошибка", "Место в центре занято")
            return

        if cargo['placed']:
            # уже размещён — удалим старые прямоугольники
            self.top_canvas.delete(cargo['top_id'])
            self.side_canvas.delete(cargo['side_id'])
            self.front_canvas.delete(cargo['front_id'])

        # создаём новые
        color = random.choice(self.colors)
        l, w, h = cargo['length'], cargo['width'], cargo['height']

        cargo['top_id'] = self.top_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE,
            fill=color, outline="black", tags=("cargo",))

        cargo['side_id'] = self.side_canvas.create_rectangle(
            10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE,
            fill=color, outline="black", tags=("cargo",))

        cargo['front_id'] = self.front_canvas.create_rectangle(
            10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
            10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE,
            fill=color, outline="black", tags=("cargo",))

        cargo['x'], cargo['y'], cargo['z'] = x, y, z
        cargo['placed'] = True
        self.update_status()

    # ------------------------------------------------------------------
    # удаление из контейнера
    # ------------------------------------------------------------------
    def remove_from_canvas(self, evt):
        canvas = evt.widget
        items = canvas.find_withtag("current")
        if not items:
            return
        clicked_id = items[0]
        cargo = None
        for c in self.cargos:
            if c.get('placed') and clicked_id in (c.get('top_id'), c.get('side_id'), c.get('front_id')):
                cargo = c
                break
        if not cargo:
            return
        # удаляем с холстов
        self.top_canvas.delete(cargo['top_id'])
        self.side_canvas.delete(cargo['side_id'])
        self.front_canvas.delete(cargo['front_id'])
        cargo['placed'] = False
        cargo.pop('top_id', None)
        cargo.pop('side_id', None)
        cargo.pop('front_id', None)
        cargo['x'] = cargo['y'] = cargo['z'] = 0
        self.update_status()

    # ------------------------------------------------------------------
    # столкновение
    # ------------------------------------------------------------------
    def collision_3d(self, x, y, z, l, w, h, exclude_id=None):
        for c in self.cargos:
            if not c.get('placed') or c['id'] == exclude_id:
                continue
            if (x < c['x'] + c['length'] and x + l > c['x'] and
                y < c['y'] + c['width']  and y + w > c['y'] and
                z < c['z'] + c['height'] and z + h > c['z']):
                return True
        return False

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def export_xlsx(self):
        if not self.cargos:
            messagebox.showwarning("Нет данных", "Нечего экспортировать")
            return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Грузы"
        ws.append(["№", "Название", "Длина (мм)", "Ширина (мм)", "Высота (мм)", "Вес (кг)"])
        for c in self.cargos:
            ws.append([c['id'], c['name'], c['length'], c['width'], c['height'], c['weight']])
        wb.save(f)
        messagebox.showinfo("Готово", f"Сохранено в {f}")

    def import_xlsx(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not f:
            return
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 6:
                    _, name, l, w, h, wt = row[:6]
                    if None in (name, l, w, h, wt):
                        continue
                    cargo = {
                        'id': self.counter,
                        'name': str(name),
                        'length': int(l), 'width': int(w), 'height': int(h),
                        'weight': int(wt),
                        'placed': False
                    }
                    self.cargos.append(cargo)
                    self.tree.insert("", "end", iid=str(self.counter),
                                     values=(name, l, w, h, wt))
                    self.counter += 1
        except Exception as e:
            messagebox.showerror("Ошибка импорта", str(e))

    # ------------------------------------------------------------------
    # drag-and-drop (только по уже размещённым)
    # ------------------------------------------------------------------
    def start_drag(self, evt):
        items = self.top_canvas.find_withtag("current")
        if not items:
            self.drag_cargo = None
            self.rotating_cargo = None
            return
        clicked_id = items[0]
        cargo = next((c for c in self.cargos
                      if c.get('placed') and c.get('top_id') == clicked_id), None)
        if cargo:
            self.drag_cargo = cargo
            self.rotating_cargo = cargo   # новое
            self.drag_start = (evt.x, evt.y)
        else:
            self.drag_cargo = None
            self.rotating_cargo = None

    def do_drag(self, evt):
        if not self.drag_cargo:
            return
        dx = int((evt.x - self.drag_start[0]) / SCALE)
        dy = int((evt.y - self.drag_start[1]) / SCALE)
        new_x = max(0, min(self.L - self.drag_cargo['length'], self.drag_cargo['x'] + dx))
        new_y = max(0, min(self.W - self.drag_cargo['width'], self.drag_cargo['y'] - dy))
        new_z = self.drag_cargo['z']
        if self.collision_3d(new_x, new_y, new_z,
                             self.drag_cargo['length'], self.drag_cargo['width'], self.drag_cargo['height'],
                             exclude_id=self.drag_cargo['id']):
            return
        self.move_cargo(new_x, new_y, new_z)
        self.drag_start = (evt.x, evt.y)

    def end_drag(self, evt):
        self.drag_cargo = None
        self.rotating_cargo = None   # груз, который может быть повернут

    def move_cargo(self, x, y, z):
        c = self.drag_cargo
        l, w, h = c['length'], c['width'], c['height']
        c['x'], c['y'], c['z'] = x, y, z
        # перерисовать
        self.top_canvas.coords(c['top_id'],
                               10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
                               10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE)
        self.side_canvas.coords(c['side_id'],
                                10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
                                10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE)
        self.front_canvas.coords(c['front_id'],
                                 10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
                                 10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE)
        self.update_status()

    # ------------------------------------------------------------------
    # misc
    # ------------------------------------------------------------------
    def update_status(self):
        placed = [c for c in self.cargos if c.get('placed')]
        total_w = sum(c['weight'] for c in placed)
        total_v = sum(c['length'] * c['width'] * c['height'] for c in placed) / 1e9
        self.status_lbl.config(text=f"Вес: {total_w} / {MAX_WEIGHT_KG} кг   "
                                    f"Объём: {total_v:.2f} / {MAX_VOLUME_M3} м³")
        
    # ------------------------------------------------------------------
    # Поворот на 90°
    # ------------------------------------------------------------------
    def on_rotate_key(self, evt):
        if not self.rotating_cargo:
            return
        c = self.rotating_cargo

        # новые размеры после поворота
        new_l = c['width']
        new_w = c['length']
        h = c['height']

        # проверим границы контейнера
        if (c['x'] + new_l > self.L) or (c['y'] + new_w > self.W):
            messagebox.showerror("Ошибка", "Поворот выходит за границы контейнера")
            return

        # проверим столкновение
        if self.collision_3d(c['x'], c['y'], c['z'],
                             new_l, new_w, h,
                             exclude_id=c['id']):
            messagebox.showerror("Ошибка", "Поворот приведёт к пересечению")
            return

        # обновляем размеры
        c['length'], c['width'] = new_l, new_w

        # перерисовываем
        self.redraw_rotated(c)
        self.update_status()

    def redraw_rotated(self, c):
        l, w, h = c['length'], c['width'], c['height']
        x, y, z = c['x'], c['y'], c['z']

        self.top_canvas.coords(c['top_id'],
                               10 + x * SCALE, 10 + (self.W - w - y) * SCALE,
                               10 + (x + l) * SCALE, 10 + (self.W - y) * SCALE)

        self.side_canvas.coords(c['side_id'],
                                10 + x * SCALE, 10 + (self.H - h - z) * SCALE,
                                10 + (x + l) * SCALE, 10 + (self.H - z) * SCALE)

        self.front_canvas.coords(c['front_id'],
                                 10 + y * SCALE, 10 + (self.H - h - z) * SCALE,
                                 10 + (y + w) * SCALE, 10 + (self.H - z) * SCALE)

    def clear_entries(self):
        for e in self.entries.values():
            e.delete(0, tk.END)


# ------------------------------------------------------------------

class Tooltip:
    """Простой tooltip для Canvas"""
    def __init__(self, canvas, app):
        self.canvas = canvas
        self.app = app               # ссылка на ContainerApp
        self.tooltip = None
        canvas.bind("<Motion>", self.on_motion)
        canvas.bind("<Leave>", self.hide)

    def on_motion(self, event):
        items = event.widget.find_withtag("current")
        if not items:
            self.hide(); return
        clicked_id = items[0]
        cargo = next((c for c in self.app.cargos
                      if c.get('placed') and clicked_id in (c.get('top_id'),
                                                            c.get('side_id'),
                                                            c.get('front_id'))), None)
        if cargo:
            text = f"{cargo['name']}  {cargo['weight']} кг"
            self.show(text, event.x_root, event.y_root)
        else:
            self.hide()

    def show(self, text, x, y):
        if self.tooltip is None:
            self.tooltip = tk.Toplevel(self.canvas)
            self.tooltip.wm_overrideredirect(True)
            self.label = tk.Label(self.tooltip, text=text, bg="yellow", relief="solid", bd=1)
            self.label.pack()
        else:
            self.label.config(text=text)
        self.tooltip.wm_geometry(f"+{x + 10}+{y + 10}")

    def hide(self, *args):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None



if __name__ == "__main__":
    root = tk.Tk()
    ContainerApp(root)
    root.mainloop()

#крайняя рабочая версия приложения. Добавлены типсы. Добавлено тримирование названий. 
#куда то пропала кнопка удалить
# добавлена возможность поворота на 90 градусов зажатием ЛКМ+R 
#ВНИМАНИЕ работает только в английской раскладке
